#!/usr/bin/env python3
# Layer: L8 — Catalyst / Meta
# Product: system-wide
# Temporal:
#   Trigger: user | cli
#   Execution: sync
# Role: Expand L2.1 intent supertable with capability candidates (no decisions)
# Callers: Claude, human operators
# Allowed Imports: L6 (pandas, openpyxl)
# Forbidden Imports: L2, L3, L4, L5
# Reference: PIN-348, Phase 2.1/2.2

"""
L2 Capability-Bound Supertable Expander

PURPOSE:
    Takes an L2.1 intent supertable and Phase-1 capability intelligence as input.
    Produces a versioned XLSX artifact with intent rows expanded into capability
    candidate rows.

    THIS IS A TRUTH EXPANDER, NOT A DECISION ENGINE.

HARD RULES:
    - Never mutate input files
    - Always generate a new versioned XLSX
    - Never select a single capability
    - Never overwrite intent columns
    - Never embed UI or product decisions
    - Never infer missing data silently
    - If logic cannot proceed → explicitly flag and continue

USAGE:
    python3 scripts/tools/l2_cap_expander.py \\
        --intent design/l2_1/supertable/L2_1_UI_INTENT_SUPERTABLE.csv \\
        --capabilities design/l2_1/elicitation/capability_intelligence_all_domains.csv \\
        --crosswalk design/l2_1/elicitation/adapter_operator_crosswalk_incidents.csv \\
        --output design/l2_1/supertable/l2_supertable_v2_cap_expanded.xlsx
"""

import argparse
import hashlib
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# =============================================================================
# CONSTANTS
# =============================================================================

VERSION = "3.0.0"
PARENT_VERSION = "2.0.0"
GENERATION_MODE = "SCRIPT"

# =============================================================================
# OVERVIEW DOMAIN DOCTRINE (v3)
# =============================================================================
# Purpose: "What's happening today? Where should I focus?"
# Nature: Cross-domain, not cross-tenant
# Mutability: None (READ-ONLY)
# Authority: None
# Source: Derived from other domain READ models
# Primary action: Navigate to underlying domain/panel
#
# Overview panels do NOT bind to platform/fdr/infra capabilities.
# They bind to a DERIVED_READ_MODEL marker indicating composition.
# =============================================================================

OVERVIEW_DOCTRINE = """
OVERVIEW DOMAIN DOCTRINE (v3)
=============================
Purpose: "What's happening today? Where should I focus?"
Nature: Cross-domain, not cross-tenant
Mutability: None (READ-ONLY)
Authority: None
Source: Derived from other domain READ models
Primary action: Navigate to underlying domain/panel

DERIVED READ MODEL CONTRACT:
- get_incidents_snapshot() → Incidents domain READ
- get_activity_snapshot() → Activity domain READ
- get_policy_snapshot() → Policies domain READ
- get_log_snapshot() → Logs domain READ

Each function returns:
- headline_metric
- trend_hint
- severity_hint
- redirect_target (domain/subdomain/topic/panel)

Overview panels do NOT bind to platform/fdr/infra capabilities (CAP-OVW-*).
They use DERIVED_READ_MODEL as a composition marker.
"""

# Overview panel to source domain mapping
OVERVIEW_SOURCE_MAPPING = {
    "System Status Summary": {
        "source_domain": "ALL",
        "operator": "get_system_status_snapshot()",
        "notes": "Aggregates health signals from all domains",
    },
    "Health Metrics Summary": {
        "source_domain": "Activity|Incidents",
        "operator": "get_health_metrics_snapshot()",
        "notes": "Derived from Activity and Incidents READ surfaces",
    },
    "Health Metrics List": {
        "source_domain": "Activity|Incidents",
        "operator": "get_health_metrics_list()",
        "notes": "O2 panel - should be O1 only per doctrine; marked for intent correction",
    },
    # Future panels per doctrine
    "Incidents Snapshot": {
        "source_domain": "Incidents",
        "operator": "get_incidents_snapshot()",
        "notes": "What needs attention now?",
    },
    "Activity Pulse": {
        "source_domain": "Activity",
        "operator": "get_activity_snapshot()",
        "notes": "What changed recently?",
    },
    "Policy Status": {
        "source_domain": "Policies",
        "operator": "get_policy_snapshot()",
        "notes": "Are controls stable?",
    },
    "Log Signals": {
        "source_domain": "Logs",
        "operator": "get_log_snapshot()",
        "notes": "Anything unusual?",
    },
}

# Required columns in intent supertable
REQUIRED_INTENT_COLUMNS = [
    "Domain",
    "Panel ID",
    "Panel Name",
    "Order",
    "Read",
    "Download",
    "Write",
    "Write Action",
    "Activate",
    "Activate Action",
]

# Required columns in capability intelligence
REQUIRED_CAP_COLUMNS = [
    "capability_id",
    "domain",
    "capability_name",
    "mode",
    "scope",
    "mutates_state",
    "bulk_support",
    "latency_profile",
    "execution_style",
    "reversibility",
    "authority_required",
    "adapters",
    "operators",
    "confidence_level",
    "l2_1_aligned",
    "l2_1_surface",
    "risk_flags",
]

# New columns to append (expansion columns)
EXPANSION_COLUMNS = [
    "row_uid",
    "action_intent",
    "candidate_capability_id",
    "candidate_operator",
    "candidate_adapter",
    "cap_risk_flags",  # Risk flags from capability intelligence
    "cap_confidence",  # Confidence level from capability intelligence
    "cap_l2_1_aligned",  # Whether capability is L2.1 aligned
    "mode_match",
    "scope_match",
    "bulk_match",
    "authority_match",
    "replay_match",
    "latency_risk",
    "binding_status",
    "binding_notes",
]

# Binding status rules
BINDING_STATUS_BLOCKED = "BLOCKED"
BINDING_STATUS_QUESTIONABLE = "QUESTIONABLE"
BINDING_STATUS_SAFE = "SAFE"


# =============================================================================
# STEP 1: LOAD & VALIDATE INPUTS
# =============================================================================


def load_intent_supertable(path: Path) -> pd.DataFrame:
    """Load intent supertable from CSV or XLSX."""
    if path.suffix.lower() == ".xlsx":
        df = pd.read_excel(path, sheet_name="SUPERTABLE")
    else:
        df = pd.read_csv(path)

    # Validate required columns
    missing = [c for c in REQUIRED_INTENT_COLUMNS if c not in df.columns]
    if missing:
        print(f"WARNING: Missing intent columns: {missing}")
        print("Proceeding with available columns...")

    print(f"Loaded {len(df)} intent rows from {path}")
    return df


def load_capability_intelligence(path: Path) -> pd.DataFrame:
    """Load capability intelligence from CSV."""
    df = pd.read_csv(path)

    # Validate required columns
    missing = [c for c in REQUIRED_CAP_COLUMNS if c not in df.columns]
    if missing:
        print(f"WARNING: Missing capability columns: {missing}")

    print(f"Loaded {len(df)} capabilities from {path}")
    return df


def load_crosswalk(path: Path | None) -> pd.DataFrame | None:
    """Load adapter/operator crosswalk if provided."""
    if path is None or not path.exists():
        print("No crosswalk file provided or found")
        return None

    df = pd.read_csv(path)
    print(f"Loaded {len(df)} crosswalk entries from {path}")
    return df


# =============================================================================
# STEP 2: NORMALIZE INTENT ROWS
# =============================================================================


def generate_row_uid(row: pd.Series) -> str:
    """Generate stable hash of intent row."""
    # Build canonical string from key fields
    parts = [
        str(row.get("Domain", "")),
        str(row.get("Panel ID", "")),
        str(row.get("Order", "")),
        str(row.get("Write Action", "")),
        str(row.get("Activate Action", "")),
    ]
    canonical = "|".join(parts)
    return hashlib.sha256(canonical.encode()).hexdigest()[:12]


def derive_action_intents(row: pd.Series) -> list[str]:
    """
    Derive action_intent from row flags.
    Returns list of intents (one row may have multiple).

    Purely mechanical — no intelligence here.
    """
    intents = []

    # READ always present if Read=YES
    if str(row.get("Read", "")).upper() == "YES":
        intents.append("READ")

    # DOWNLOAD if Download=YES
    if str(row.get("Download", "")).upper() == "YES":
        intents.append("DOWNLOAD")

    # WRITE:<action> if Write=YES and Write Action exists
    if str(row.get("Write", "")).upper() == "YES":
        write_action = str(row.get("Write Action", "")).strip()
        if write_action and write_action.upper() != "NAN":
            intents.append(f"WRITE:{write_action}")
        else:
            intents.append("WRITE:UNKNOWN")

    # ACTIVATE:<action> if Activate=YES and Activate Action exists
    if str(row.get("Activate", "")).upper() == "YES":
        activate_action = str(row.get("Activate Action", "")).strip()
        if activate_action and activate_action.upper() != "NAN":
            # Handle pipe-separated actions
            for action in activate_action.split("|"):
                intents.append(f"ACTIVATE:{action.strip()}")
        else:
            intents.append("ACTIVATE:UNKNOWN")

    return intents if intents else ["READ"]  # Default to READ if nothing specified


# =============================================================================
# STEP 3: CANDIDATE CAPABILITY SEARCH
# =============================================================================


def normalize_domain(domain: str) -> str:
    """Normalize domain name for matching."""
    return str(domain).strip().lower()


def normalize_surface(surface: str) -> str:
    """Normalize L2.1 surface for matching."""
    return str(surface).strip().upper()


def get_mode_from_intent(intent: str) -> str:
    """Extract mode from action intent."""
    if intent.startswith("WRITE:") or intent.startswith("ACTIVATE:"):
        return "WRITE"
    return "READ"


def find_candidate_capabilities(
    intent: str,
    row: pd.Series,
    capabilities_df: pd.DataFrame,
) -> list[dict[str, Any]]:
    """
    Find candidate capabilities for an action intent.
    Returns list of candidate dicts with match flags.

    NO SCORING. NO RANKING. Just evidence.
    """
    candidates = []

    # Extract domain and surface from intent row
    row_domain = normalize_domain(row.get("Domain", ""))
    row_topic_id = str(row.get("Topic ID", ""))

    # Determine expected mode
    expected_mode = get_mode_from_intent(intent)

    # Search capabilities
    for _, cap in capabilities_df.iterrows():
        cap_domain = normalize_domain(cap.get("domain", ""))
        cap_mode = str(cap.get("mode", "")).upper()
        cap_surface = normalize_surface(cap.get("l2_1_surface", ""))
        cap_id = str(cap.get("capability_id", ""))

        # Domain match check
        domain_match = cap_domain == row_domain

        # Surface match check (flexible)
        surface_match = False
        if cap_surface and cap_surface != "NAN":
            if row_topic_id and row_topic_id in cap_surface:
                surface_match = True
            elif row_domain.upper() in cap_surface:
                surface_match = True

        # Mode match check
        mode_match = cap_mode == expected_mode

        # Only include if domain matches or surface matches
        if domain_match or surface_match:
            # Extract risk flags and confidence from capability
            risk_flags = str(cap.get("risk_flags", ""))
            confidence = str(cap.get("confidence_level", ""))
            l2_1_aligned = str(cap.get("l2_1_aligned", ""))

            # Build candidate record
            candidate = {
                "candidate_capability_id": cap_id,
                "candidate_operator": str(cap.get("operators", "")),
                "candidate_adapter": str(cap.get("adapters", "")),
                "cap_risk_flags": risk_flags
                if risk_flags and risk_flags != "nan"
                else "",
                "cap_confidence": confidence,
                "cap_l2_1_aligned": l2_1_aligned,
                "mode_match": "YES" if mode_match else "NO",
                "scope_match": "YES" if surface_match else "PARTIAL",
                "bulk_match": evaluate_bulk_match(intent, cap),
                "authority_match": evaluate_authority_match(cap),
                "replay_match": evaluate_replay_match(row, cap),
                "latency_risk": evaluate_latency_risk(cap),
            }
            candidates.append(candidate)

    return candidates


def evaluate_bulk_match(intent: str, cap: pd.Series) -> str:
    """Evaluate bulk operation compatibility."""
    bulk_support = str(cap.get("bulk_support", "")).upper()
    scope = str(cap.get("scope", "")).upper()

    if bulk_support == "YES" or scope == "BULK":
        return "YES"
    elif bulk_support == "NO" or scope == "SINGLE":
        return "NO"
    return "UNKNOWN"


def evaluate_authority_match(cap: pd.Series) -> str:
    """Evaluate authority/permission compatibility."""
    authority = str(cap.get("authority_required", "")).upper()
    mutates = str(cap.get("mutates_state", "")).upper()

    if authority in ["NONE", ""] or authority == "NAN":
        return "YES"
    elif mutates == "YES":
        return "REQUIRES_AUTH"
    return "UNKNOWN"


def evaluate_replay_match(row: pd.Series, cap: pd.Series) -> str:
    """Evaluate replay compatibility."""
    row_replay = str(row.get("Replay", "")).upper()
    cap_mutates = str(cap.get("mutates_state", "")).upper()

    if row_replay == "YES" and cap_mutates == "YES":
        return "CONFLICT"  # Replay + mutation = potential issue
    elif row_replay == "YES":
        return "YES"
    return "N/A"


def evaluate_latency_risk(cap: pd.Series) -> str:
    """Evaluate latency risk from capability profile."""
    latency = str(cap.get("latency_profile", "")).upper()

    if latency in ["HIGH", "MEDIUM-HIGH"]:
        return "HIGH"
    elif latency in ["MEDIUM"]:
        return "MEDIUM"
    elif latency in ["LOW", "VERY LOW"]:
        return "LOW"
    return "UNKNOWN"


# =============================================================================
# STEP 4: EXPAND ROWS
# =============================================================================


def expand_overview_row(row: pd.Series, row_uid: str) -> list[dict[str, Any]]:
    """
    Expand an Overview domain row using DERIVED_READ_MODEL binding.

    Overview panels do NOT bind to platform/founder capabilities (CAP-OVW-*).
    They bind to a composition marker indicating derived data from other domains.

    Returns list with single expanded row (one-to-one for Overview).
    """
    original_data = row.to_dict()
    panel_name = str(row.get("Panel Name", ""))
    order = str(row.get("Order", ""))

    # Get mapping for this panel
    mapping = OVERVIEW_SOURCE_MAPPING.get(
        panel_name,
        {
            "source_domain": "UNKNOWN",
            "operator": "get_overview_snapshot()",
            "notes": "Panel not in doctrine mapping",
        },
    )

    # Build binding notes
    notes_parts = [
        f"Derived from {mapping['source_domain']} domain READ capabilities",
        "No direct platform binding",
    ]

    # Check for O2 violation (Overview should be O1 only per doctrine)
    if order != "O1":
        notes_parts.append(
            f"INTENT_CORRECTION_NEEDED: {order} should be O1 per doctrine"
        )

    expanded_row = original_data.copy()
    expanded_row.update(
        {
            "row_uid": row_uid,
            "action_intent": "READ",
            "candidate_capability_id": "DERIVED_READ_MODEL",
            "candidate_operator": f"customer_console_overview.py::{mapping['operator']}",
            "candidate_adapter": "LOCAL_COMPOSITION",
            "cap_risk_flags": "",
            "cap_confidence": "HIGH",
            "cap_l2_1_aligned": "YES",
            "mode_match": "YES",
            "scope_match": "YES",
            "bulk_match": "N/A",
            "authority_match": "YES",
            "replay_match": "N/A",
            "latency_risk": "LOW",
            "binding_status": BINDING_STATUS_SAFE,
            "binding_notes": "; ".join(notes_parts),
        }
    )

    return [expanded_row]


def expand_intent_row(
    row: pd.Series,
    capabilities_df: pd.DataFrame,
    crosswalk_df: pd.DataFrame | None,
) -> list[dict[str, Any]]:
    """
    Expand a single intent row into multiple capability-candidate rows.
    Returns list of expanded row dicts.
    """
    expanded = []

    # Generate row UID
    row_uid = generate_row_uid(row)

    # OVERVIEW DOMAIN SPECIAL HANDLING (v3)
    # Overview panels bind to DERIVED_READ_MODEL, not platform capabilities
    domain = str(row.get("Domain", "")).strip()
    if domain.upper() == "OVERVIEW":
        return expand_overview_row(row, row_uid)

    # Get all action intents
    action_intents = derive_action_intents(row)

    # Convert original row to dict (preserve all columns)
    original_data = row.to_dict()

    for intent in action_intents:
        # Find candidates for this intent
        candidates = find_candidate_capabilities(intent, row, capabilities_df)

        if not candidates:
            # No candidates found — create BLOCKED row
            expanded_row = original_data.copy()
            expanded_row.update(
                {
                    "row_uid": row_uid,
                    "action_intent": intent,
                    "candidate_capability_id": "NONE",
                    "candidate_operator": "N/A",
                    "candidate_adapter": "N/A",
                    "mode_match": "N/A",
                    "scope_match": "N/A",
                    "bulk_match": "N/A",
                    "authority_match": "N/A",
                    "replay_match": "N/A",
                    "latency_risk": "N/A",
                    "binding_status": BINDING_STATUS_BLOCKED,
                    "binding_notes": "No candidate capabilities found for this intent",
                }
            )
            expanded.append(expanded_row)
        else:
            # Create one row per candidate
            for candidate in candidates:
                expanded_row = original_data.copy()
                expanded_row["row_uid"] = row_uid
                expanded_row["action_intent"] = intent
                expanded_row.update(candidate)

                # Enrich from crosswalk if available
                if crosswalk_df is not None:
                    enrichment = enrich_from_crosswalk(
                        candidate["candidate_capability_id"],
                        crosswalk_df,
                    )
                    if enrichment:
                        expanded_row.update(enrichment)

                # Classify binding status
                binding_status, notes = classify_binding_status(expanded_row)
                expanded_row["binding_status"] = binding_status
                expanded_row["binding_notes"] = notes

                expanded.append(expanded_row)

    return expanded


def enrich_from_crosswalk(
    capability_id: str,
    crosswalk_df: pd.DataFrame,
) -> dict[str, Any] | None:
    """Enrich candidate with crosswalk information."""
    matches = crosswalk_df[crosswalk_df["capability_id"] == capability_id]
    if matches.empty:
        return None

    # Take first match (could be multiple operators)
    match = matches.iloc[0]
    return {
        "crosswalk_adapter": str(match.get("adapter_id", "")),
        "crosswalk_operator": str(match.get("operator_name", "")),
        "crosswalk_layer_route": str(match.get("layer_route", "")),
        "crosswalk_side_effects": str(match.get("side_effects", "")),
    }


# =============================================================================
# STEP 5: BINDING STATUS CLASSIFICATION (RULE-BASED)
# =============================================================================


def classify_binding_status(row: dict[str, Any]) -> tuple[str, str]:
    """
    Classify binding status using explicit rules.
    Returns (status, notes).

    RULES (explicit and logged):
    - BLOCKED: No candidates
    - QUESTIONABLE: Any critical mismatch or risk flag
    - SAFE: All matches pass
    """
    notes_parts = []

    # Check for no candidates
    if row.get("candidate_capability_id") == "NONE":
        return BINDING_STATUS_BLOCKED, "No candidate capabilities"

    # Check for critical mismatches
    critical_issues = []

    # Mode mismatch is critical
    if row.get("mode_match") == "NO":
        critical_issues.append("Mode mismatch (READ vs WRITE)")

    # Authority issues
    if row.get("authority_match") == "REQUIRES_AUTH":
        notes_parts.append("Requires human authorization")

    # Replay conflicts
    if row.get("replay_match") == "CONFLICT":
        critical_issues.append("Replay/mutation conflict")

    # High latency
    if row.get("latency_risk") == "HIGH":
        notes_parts.append("High latency risk")

    # Check risk flags from capability intelligence
    risk_flags = str(row.get("cap_risk_flags", ""))
    if risk_flags and risk_flags != "nan":
        # Critical risk flags that indicate capability is not ready
        critical_flags = [
            "NO Customer API",
            "NOT IMPLEMENTED",
            "ADAPTER BYPASSED",
            "Founder-only",
            "REVERSIBILITY UNKNOWN",
            "IDEMPOTENCY UNCLEAR",
        ]
        for flag in critical_flags:
            if flag.upper() in risk_flags.upper():
                critical_issues.append(f"Risk flag: {flag}")
                break  # Only add one risk flag note

    # Check L2.1 alignment
    l2_1_aligned = str(row.get("cap_l2_1_aligned", "")).upper()
    if l2_1_aligned == "NO":
        critical_issues.append("Not L2.1 aligned")
    elif l2_1_aligned == "PARTIAL":
        notes_parts.append("Partial L2.1 alignment")

    # Determine status
    if critical_issues:
        notes = "; ".join(critical_issues + notes_parts)
        return BINDING_STATUS_QUESTIONABLE, notes

    if notes_parts:
        return BINDING_STATUS_SAFE, "; ".join(notes_parts)

    return BINDING_STATUS_SAFE, "All checks pass"


# =============================================================================
# STEP 6: WRITE XLSX OUTPUT
# =============================================================================


def create_changelog_df(
    intent_df: pd.DataFrame,
    expanded_df: pd.DataFrame,
    original_columns: list[str],
) -> pd.DataFrame:
    """Create changelog dataframe."""
    columns_added = [c for c in EXPANSION_COLUMNS if c not in original_columns]

    # Count binding statuses
    status_counts = expanded_df["binding_status"].value_counts().to_dict()

    # Count Overview rows
    overview_count = len(expanded_df[expanded_df["Domain"] == "Overview"])

    # Build v3 summary
    if VERSION.startswith("3"):
        summary = (
            "v3: Overview domain corrected to derived, cross-domain customer snapshot; "
            "removed platform health bindings (CAP-OVW-*); "
            "Overview now uses DERIVED_READ_MODEL marker"
        )
        domains_affected = "Overview"
        overview_doctrine_note = OVERVIEW_DOCTRINE.strip()
    else:
        summary = "Expanded intent rows with capability candidates"
        domains_affected = "All"
        overview_doctrine_note = "N/A"

    changelog = {
        "version": VERSION,
        "parent_version": PARENT_VERSION,
        "generation_mode": GENERATION_MODE,
        "date": datetime.now().isoformat(),
        "summary_of_change": summary,
        "domains_affected": domains_affected,
        "rows_in_parent": len(intent_df),
        "rows_in_this_version": len(expanded_df),
        "overview_rows_modified": overview_count,
        "columns_added": ", ".join(columns_added),
        "columns_removed": "None",
        "known_limitations": "; ".join(
            [
                "Crosswalk limited to Incidents domain",
                "Activity/Logs/Policies crosswalks not provided",
            ]
        ),
        "review_status": "DRAFT",
        "binding_stats_safe": status_counts.get(BINDING_STATUS_SAFE, 0),
        "binding_stats_questionable": status_counts.get(BINDING_STATUS_QUESTIONABLE, 0),
        "binding_stats_blocked": status_counts.get(BINDING_STATUS_BLOCKED, 0),
        "overview_doctrine": overview_doctrine_note,
    }

    return pd.DataFrame([changelog])


def write_xlsx_output(
    changelog_df: pd.DataFrame,
    expanded_df: pd.DataFrame,
    output_path: Path,
) -> None:
    """Write output XLSX with CHANGELOG and SUPERTABLE sheets."""
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create CHANGELOG sheet
    ws_changelog = wb.create_sheet("CHANGELOG")

    # Style for headers
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    # Write changelog
    for r_idx, row in enumerate(
        dataframe_to_rows(changelog_df.T.reset_index(), index=False, header=False)
    ):
        for c_idx, value in enumerate(row, 1):
            cell = ws_changelog.cell(row=r_idx + 1, column=c_idx, value=value)
            if c_idx == 1:
                cell.font = Font(bold=True)

    ws_changelog.column_dimensions["A"].width = 25
    ws_changelog.column_dimensions["B"].width = 80

    # Create SUPERTABLE sheet
    ws_super = wb.create_sheet("SUPERTABLE")

    for r_idx, row in enumerate(
        dataframe_to_rows(expanded_df, index=False, header=True), 1
    ):
        for c_idx, value in enumerate(row, 1):
            cell = ws_super.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Auto-width columns (limited)
    for col_idx, col in enumerate(expanded_df.columns, 1):
        ws_super.column_dimensions[ws_super.cell(1, col_idx).column_letter].width = min(
            max(len(str(col)) + 2, 12), 40
        )

    # Save
    wb.save(output_path)
    print(f"Wrote {output_path}")


def write_blocked_rows_csv(expanded_df: pd.DataFrame, output_dir: Path) -> None:
    """Write blocked rows to separate CSV."""
    blocked = expanded_df[expanded_df["binding_status"] == BINDING_STATUS_BLOCKED]
    if not blocked.empty:
        path = output_dir / "blocked_rows.csv"
        blocked.to_csv(path, index=False)
        print(f"Wrote {len(blocked)} blocked rows to {path}")


def write_multi_candidate_csv(expanded_df: pd.DataFrame, output_dir: Path) -> None:
    """Write rows with multiple candidates to separate CSV."""
    # Group by row_uid and action_intent, count candidates
    grouped = expanded_df.groupby(["row_uid", "action_intent"]).size()
    multi = grouped[grouped > 1]

    if not multi.empty:
        # Get all rows for these groups
        multi_rows = expanded_df[
            expanded_df.set_index(["row_uid", "action_intent"]).index.isin(multi.index)
        ]
        path = output_dir / "multi_candidate_rows.csv"
        multi_rows.to_csv(path, index=False)
        print(f"Wrote {len(multi_rows)} multi-candidate rows to {path}")


# =============================================================================
# MAIN
# =============================================================================


def print_summary(intent_df: pd.DataFrame, expanded_df: pd.DataFrame) -> None:
    """Print console summary."""
    print("\n" + "=" * 60)
    print("L2 CAPABILITY EXPANDER SUMMARY")
    print("=" * 60)
    print(f"  Total intent rows:     {len(intent_df)}")
    print(f"  Total expanded rows:   {len(expanded_df)}")
    print()

    status_counts = expanded_df["binding_status"].value_counts()
    print("  Binding Status:")
    for status, count in status_counts.items():
        pct = (count / len(expanded_df)) * 100
        print(f"    {status:15s}: {count:4d} ({pct:5.1f}%)")

    print()

    # Unique intents
    intent_counts = expanded_df["action_intent"].value_counts()
    print("  Action Intents:")
    for intent, count in intent_counts.head(10).items():
        print(f"    {intent:30s}: {count}")

    print("=" * 60)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="L2 Capability-Bound Supertable Expander",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--intent",
        type=Path,
        required=True,
        help="Path to L2.1 intent supertable (CSV or XLSX)",
    )
    parser.add_argument(
        "--capabilities",
        type=Path,
        required=True,
        help="Path to capability intelligence CSV",
    )
    parser.add_argument(
        "--crosswalk",
        type=Path,
        default=None,
        help="Path to adapter/operator crosswalk CSV (optional)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        required=True,
        help="Output path for expanded XLSX",
    )
    parser.add_argument(
        "--export-csv",
        action="store_true",
        help="Also export blocked_rows.csv and multi_candidate_rows.csv",
    )

    args = parser.parse_args()

    # Validate inputs exist
    if not args.intent.exists():
        print(f"ERROR: Intent file not found: {args.intent}")
        return 1

    if not args.capabilities.exists():
        print(f"ERROR: Capabilities file not found: {args.capabilities}")
        return 1

    # STEP 1: Load inputs
    print("STEP 1: Loading inputs...")
    intent_df = load_intent_supertable(args.intent)
    capabilities_df = load_capability_intelligence(args.capabilities)
    crosswalk_df = load_crosswalk(args.crosswalk)

    # Preserve original columns
    original_columns = list(intent_df.columns)

    # STEP 2-4: Expand rows
    print("\nSTEP 2-4: Normalizing and expanding rows...")
    expanded_rows = []
    for idx, row in intent_df.iterrows():
        expanded = expand_intent_row(row, capabilities_df, crosswalk_df)
        expanded_rows.extend(expanded)
        if (idx + 1) % 10 == 0:
            print(f"  Processed {idx + 1}/{len(intent_df)} intent rows...")

    expanded_df = pd.DataFrame(expanded_rows)

    # Reorder columns: original first, then expansion
    final_columns = original_columns + [
        c for c in EXPANSION_COLUMNS if c in expanded_df.columns
    ]
    # Add any crosswalk columns
    crosswalk_cols = [c for c in expanded_df.columns if c.startswith("crosswalk_")]
    final_columns.extend(crosswalk_cols)
    expanded_df = expanded_df[[c for c in final_columns if c in expanded_df.columns]]

    # STEP 5: Already done in expand_intent_row

    # STEP 6: Write output
    print("\nSTEP 6: Writing output...")
    changelog_df = create_changelog_df(intent_df, expanded_df, original_columns)

    # Ensure output directory exists
    args.output.parent.mkdir(parents=True, exist_ok=True)

    write_xlsx_output(changelog_df, expanded_df, args.output)

    # Optional CSV exports
    if args.export_csv:
        write_blocked_rows_csv(expanded_df, args.output.parent)
        write_multi_candidate_csv(expanded_df, args.output.parent)

    # Print summary
    print_summary(intent_df, expanded_df)

    return 0


if __name__ == "__main__":
    sys.exit(main())
